from __future__ import annotations

import argparse
import csv
import html
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DART = "https://opendart.fss.or.kr/api/list.json"
NXT_PAGE = "https://www.nextrade.co.kr/menu/marketData/menuList.do"
NXT_API = "https://www.nextrade.co.kr/brdinfoTime/brdinfoTimeList.do"
NAVER_QUOTE = "https://polling.finance.naver.com/api/realtime"
KEYWORDS = ("계약", "공급", "수주", "단일판매", "유상증자", "무상증자", "합병", "인수", "특허", "승인", "허가", "투자", "자금조달", "최대주주")
UPPER_LIMIT_THRESHOLD = 29.0


def code(value: Any) -> str:
    s = re.sub(r"\D", "", str(value or ""))
    return s[-6:].zfill(6) if s else ""


def num(value: Any) -> float | None:
    try:
        return float(str(value).replace(",", ""))
    except Exception:
        return None


def fetch_nxt() -> dict[str, str]:
    s = requests.Session()
    headers = {"User-Agent": "Mozilla/5.0", "Referer": NXT_PAGE}
    s.get(NXT_PAGE, headers=headers, timeout=15)
    r = s.post(NXT_API, data={"pageIndex": 1, "pageUnit": 2000}, headers=headers, timeout=20)
    r.raise_for_status()
    rows = r.json().get("brdinfoTimeList", [])
    out: dict[str, str] = {}
    for row in rows:
        raw = str(row.get("isuSrdCd") or "")
        c = code(raw[1:] if raw.startswith("A") else raw)
        if c:
            out[c] = row.get("isuAbwdNm") or ""
    return out


def fetch_dart(day: str, nxt: dict[str, str]) -> list[dict[str, Any]]:
    key = os.getenv("DART_API_KEY") or os.getenv("API_K_DART")
    if not key:
        raise SystemExit("DART_API_KEY secret is missing")
    ymd = day.replace("-", "")
    rows: list[dict[str, Any]] = []
    page = 1
    while True:
        params = {"crtfc_key": key, "bgn_de": ymd, "end_de": ymd, "page_no": page, "page_count": 100}
        data = requests.get(DART, params=params, timeout=20).json()
        if data.get("status") not in ("000", "013"):
            raise RuntimeError(f"DART error {data.get('status')}: {data.get('message')}")
        batch = data.get("list") or []
        for row in batch:
            c = code(row.get("stock_code"))
            if c in nxt:
                title = row.get("report_nm") or ""
                rows.append({
                    "stock_code": c,
                    "stock_name": row.get("corp_name") or nxt.get(c, ""),
                    "disclosure_time": row.get("rcept_dt") or ymd,
                    "disclosure_title": title,
                    "dart_url": f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={row.get('rcept_no')}",
                    "reason": "",
                    "source": "",
                    "score": 10 + sum(8 for k in KEYWORDS if k in title),
                })
        if page >= int(data.get("total_page") or 1):
            break
        page += 1
    return rows


def fetch_quotes(codes: list[str]) -> dict[str, dict[str, Any]]:
    s = requests.Session()
    headers = {"User-Agent": "Mozilla/5.0", "Referer": "https://finance.naver.com/"}
    quotes: dict[str, dict[str, Any]] = {}
    for c in codes:
        try:
            r = s.get(NAVER_QUOTE, params={"query": f"SERVICE_ITEM:{c}"}, headers=headers, timeout=8)
            datas = (r.json().get("result", {}).get("areas") or [{}])[0].get("datas") or []
        except Exception:
            datas = []
        for row in datas:
            rc = code(row.get("cd"))
            if rc:
                quotes[rc] = {
                    "stock_code": rc,
                    "stock_name": row.get("nm") or "",
                    "close": num(row.get("nv")),
                    "change": num(row.get("cv")),
                    "change_rate": num(row.get("cr")) or 0,
                    "volume": num(row.get("aq")),
                    "trading_value": num(row.get("aa")),
                }
    return quotes


def fetch_latest_upper_limit_stocks(nxt: dict[str, str]) -> list[dict[str, Any]]:
    try:
        import FinanceDataReader as fdr
    except ImportError as exc:
        raise SystemExit("FinanceDataReader is required for morning upper-limit scan") from exc

    frames = []
    for market in ("KOSPI", "KOSDAQ"):
        frame = fdr.StockListing(market)
        if frame is not None and not frame.empty:
            frames.append(frame)
    if not frames:
        return []

    rows: list[dict[str, Any]] = []
    for frame in frames:
        for _, item in frame.iterrows():
            c = code(item.get("Code"))
            if c not in nxt:
                continue
            change_rate = num(item.get("ChagesRatio")) or 0
            if change_rate < UPPER_LIMIT_THRESHOLD:
                continue
            rows.append(
                {
                    "stock_code": c,
                    "stock_name": item.get("Name") or nxt.get(c, ""),
                    "close": num(item.get("Close")),
                    "change": num(item.get("Changes")),
                    "change_rate": change_rate,
                    "volume": num(item.get("Volume")),
                    "trading_value": num(item.get("Amount")),
                }
            )
    return sorted(rows, key=lambda x: x.get("change_rate") or 0, reverse=True)


def news_reason(name: str, c: str) -> tuple[str, str]:
    q = f"{name} {c} 급등 이유"
    url = "https://search.naver.com/search.naver"
    try:
        text = requests.get(url, params={"where": "news", "query": q}, headers={"User-Agent": "Mozilla/5.0"}, timeout=10).text
        soup = BeautifulSoup(text, "html.parser")
        item = soup.select_one("a.news_tit") or soup.select_one("a[href*='news']")
        if not item:
            return "", ""
        title = item.get("title") or item.get_text(" ", strip=True)
        return title[:120], item.get("href") or ""
    except Exception:
        return "", ""


def merge_rows(disclosures: list[dict[str, Any]], quotes: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    for d in disclosures:
        rows[d["stock_code"]] = d
    for c, q in quotes.items():
        if (q.get("change_rate") or 0) >= 25:
            rows.setdefault(c, {
                "stock_code": c,
                "stock_name": q.get("stock_name", ""),
                "disclosure_time": "",
                "disclosure_title": "당일 +25% 이상 상승",
                "dart_url": "",
                "reason": "",
                "source": "",
                "score": 20,
            })
    for c, row in rows.items():
        q = quotes.get(c, {})
        row.update({k: q.get(k) for k in ("close", "change", "change_rate", "volume", "trading_value")})
        row["stock_name"] = row.get("stock_name") or q.get("stock_name", "")
        if (row.get("change_rate") or 0) >= 28:
            row["reason"], row["source"] = news_reason(row["stock_name"], c)
    return sorted(rows.values(), key=lambda x: (x.get("change_rate") or 0, x.get("score") or 0), reverse=True)


def save_reports(rows: list[dict[str, Any]], day: str) -> None:
    out = Path("reports")
    out.mkdir(exist_ok=True)
    cols = ["stock_code", "stock_name", "disclosure_time", "close", "change", "change_rate", "volume", "trading_value", "disclosure_title", "reason", "source", "dart_url", "score"]
    with (out / f"{day}_nxt_candidates.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    (out / f"{day}_nxt_candidates.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    body = "".join("<tr>" + "".join(f"<td>{html.escape(str(r.get(c, '') or ''))}</td>" for c in cols) + "</tr>" for r in rows)
    (out / f"{day}_nxt_report.html").write_text(f"<html><body><table>{body}</table></body></html>", encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "NXT candidates"
    headers = ["종목코드", "종목명", "공시시간", "종가", "전일대비", "등락률", "거래량", "거래대금", "공시/이벤트", "상승 이유", "출처", "DART", "점수"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    hot = PatternFill("solid", fgColor="FFE699")
    for r in rows:
        ws.append([r.get(c) for c in cols])
        if (r.get("change_rate") or 0) >= 28:
            for cell in ws[ws.max_row]:
                cell.fill = hot
    for i, width in enumerate([12, 18, 14, 12, 12, 10, 14, 16, 46, 46, 32, 48, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    wb.save(out / f"{day}_nxt_candidates.xlsx")


def telegram_enabled() -> bool:
    return bool(os.getenv("TELEGRAM_BOT_TOKEN") and os.getenv("TELEGRAM_CHAT_ID"))


def send_telegram(text: str) -> None:
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        print("Telegram secrets are missing; skipping Telegram send.")
        return
    response = requests.post(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data={
            "chat_id": chat_id,
            "text": text,
            "disable_web_page_preview": "true",
        },
        timeout=15,
    )
    response.raise_for_status()


def format_report_message(rows: list[dict[str, Any]], day: str) -> str:
    run_url = os.getenv("GITHUB_RUN_URL", "")
    hot = [row for row in rows if (row.get("change_rate") or 0) >= 28]
    lines = [
        f"[NXT 공시/급등 스캐너] {day}",
        f"후보: {len(rows)}개 / +28% 이상: {len(hot)}개",
    ]
    for row in rows[:10]:
        rate = row.get("change_rate")
        rate_text = f"{rate:.2f}%" if isinstance(rate, (int, float)) else "-"
        lines.append(f"- {row.get('stock_name')}({row.get('stock_code')}): {rate_text} / {row.get('disclosure_title')}")
    if run_url:
        lines.append(f"결과 다운로드: {run_url}")
    return "\n".join(lines)


def format_upper_limit_message(rows: list[dict[str, Any]]) -> str:
    today = datetime.now().strftime("%Y-%m-%d")
    lines = [f"[전일 상한가 NXT 종목] {today} 07:00 점검"]
    if not rows:
        lines.append("전일 상한가권(+29% 이상) NXT 종목 없음")
        return "\n".join(lines)
    for row in rows:
        rate = row.get("change_rate")
        close = row.get("close")
        rate_text = f"{rate:.2f}%" if isinstance(rate, (int, float)) else "-"
        close_text = f"{close:,.0f}원" if isinstance(close, (int, float)) else "-"
        lines.append(f"- {row.get('stock_name')}({row.get('stock_code')}): {rate_text}, 종가 {close_text}")
    return "\n".join(lines)


def run_morning_upper_limit() -> None:
    nxt = fetch_nxt()
    rows = fetch_latest_upper_limit_stocks(nxt)
    out = Path("reports")
    out.mkdir(exist_ok=True)
    day = datetime.now().strftime("%Y-%m-%d")
    (out / f"{day}_morning_upper_limit_nxt.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    send_telegram(format_upper_limit_message(rows))
    print(f"NXT={len(nxt)} upper_limit_nxt={len(rows)}")


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"))
    p.add_argument("--mode", choices=("daily-report", "morning-upper-limit"), default="daily-report")
    args = p.parse_args()
    if args.mode == "morning-upper-limit":
        run_morning_upper_limit()
        return

    nxt = fetch_nxt()
    disclosures = fetch_dart(args.date, nxt)
    filtered = [d for d in disclosures if any(k in d["disclosure_title"] for k in KEYWORDS)]
    quotes = fetch_quotes(sorted(set(nxt) | {d["stock_code"] for d in filtered}))
    rows = merge_rows(filtered, quotes)
    save_reports(rows, args.date)
    if telegram_enabled():
        send_telegram(format_report_message(rows, args.date))
    print(f"NXT={len(nxt)} disclosures={len(disclosures)} candidates={len(rows)}")


if __name__ == "__main__":
    main()
